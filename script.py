import csv
import api_requests
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# roles
# volunteers

#steps:
# 1. model volunteer and their availability
# 2. model roles and their scheduling needs
# 3. print both models to csv

class Volunteer:
    def __init__(self, name, availability):
        self.name = name
        self.availability = availability

    def __str__(self):
        return f'{self.name}, {self.availability}'

class Role: 
    def __init__(self, name, schedule):
        self.name = name
        self.schedule = schedule

    def __str__(self):
        return f'{self.name}, {self.schedule}'


if __name__ == "__main__":
    with open('data.json') as f:
        data = json.load(f)
        
        volunteers = []

        for response in data["responses"]:
            values = response["values"]
            name = values["QID1_TEXT"] + " " + values["QID2_TEXT"]

            labels = response["labels"]
            try:
                availability = [labels["QID17"], labels["QID18"]]
            except KeyError:
                availability = [labels["QID13"], labels["QID14"]]
            
            availability = [availability[0], availability[1]]

            a_temp = []

            for a in availability:

                # Regular expression pattern to match times
                pattern = r'(\d{1,2}:\d{2}[APM]{2})\s*-\s*(\d{1,2}:\d{2}[APM]{2})'

                # Search for the pattern in the string
                match = re.search(pattern, a)

                if match:
                    start_time = match.group(1)
                    end_time = match.group(2)

                start_int = int(start_time.split(":")[0])
                end_int = int(end_time.split(":")[0])

                for i in range(start_int, end_int):
                    a_temp.append(i)

            availability = a_temp

            V = Volunteer(name, availability)
            volunteers.append(V)
            
def to_csv(volunteers):
    with open('volunteers.csv', mode='w') as file:
        writer = csv.writer(file)
        writer.writerow(['Name', 'Availability'])
        for volunteer in volunteers:
            writer.writerow([volunteer.name, volunteer.availability])


def colourize(csv_file):
    # Load CSV file into a DataFrame
    df = pd.read_csv(csv_file)

    # Save DataFrame to an Excel file
    excel_file = csv_file.replace('.csv', '.xlsx')
    df.to_excel(excel_file, index=False)

    # Load the Excel file with openpyxl
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Define a fill pattern (color)
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green fill
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill

    # Apply the fill color to specific cells
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            # Apply color conditionally or to all cells
            if cell.value == "Specific Value":  # Replace with your condition
                cell.fill = fill_color

    # Save the modified Excel file
    workbook.save('colored_file.xlsx')